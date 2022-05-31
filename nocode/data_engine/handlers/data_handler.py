# -*- coding: utf-8 -*-
"""
Tencent is pleased to support the open source community by making BK-NOCODE SMAKER蓝鲸无代码平台  available.

Copyright (C) 2021 THL A29 Limited, a Tencent company.  All rights reserved.

BK-NOCODE 蓝鲸无代码平台(S-maker) is licensed under the MIT License.

License for BK-NOCODE 蓝鲸无代码平台(S-maker) :
--------------------------------------------------------------------
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software,
and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial
portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT
LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN
NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""
import ast
import datetime
import io
import json
import re

import openpyxl
import xlwt
from django.db.models import Count
from django.http import HttpResponse
from rest_framework.response import Response

from common.log import logger
from itsm.component.drf.pagination import CustomPageNumberPagination
from nocode.base.constants import SELECT_FIELDS_TYPE
from nocode.data_engine.core.constants import DAY, MONTH, YEAR
from nocode.data_engine.core.managers import DataManager

from nocode.data_engine.core.utils import (
    ConditionTransfer,
    compute_time_range,
    value_to_list,
)
from nocode.data_engine.exceptions import (
    GetDetailDataError,
    ImportDataError,
    DataValidateError,
)
from nocode.data_engine.handlers.constancts import CONDITIONS_PROTOCOL, AUTO_LINE_FEED
from nocode.data_engine.handlers.module_handlers import (
    PageComponentHandler,
    ServiceHandler,
)
from nocode.project_manager.models import ProjectVersion


class BaseDataHandler:
    def get_response_data(self, page_queryset):
        """
        查询到的值去掉contents__前缀
        """
        data = []
        for item in page_queryset:
            new_item = {}
            for key, value in item.items():
                new_item[key.replace("contents__", "", 1)] = value
            data.append(new_item)

        return data

    def convert_conditions(self, conditions):
        transfer = ConditionTransfer(conditions)
        return transfer.trans()

    def get_config(self, config):
        # 对之前对数据做兼容
        if isinstance(config, str):
            return json.loads(config)
        return config


class ListComponentDataHandler(BaseDataHandler):
    def __init__(self, page_id, request, version_number):
        self.request = request
        self.page_id = page_id
        self.version_number = version_number

    def get_list_components_data(self, conditions=None, need_page=True, tab_id=None):
        """
        获取列表组件的data
        """
        if conditions is None:
            conditions = {}

        filters = self.convert_conditions(conditions)

        page_config = PageComponentHandler.get_list_component_config(
            self.page_id, self.version_number
        )
        worksheet_id = page_config["value"]

        logger.info(
            "[get_list_components_data] -> worksheet_id={}".format(worksheet_id)
        )

        manager = DataManager(worksheet_id)
        fields = manager.fields
        keys, upload_fields = self.get_keys(page_config["config"], fields)
        logger.info("[get_list_components_data] -> keys={}".format(keys))

        page_config_queryset = self.get_page_config_queryset(
            manager, page_config["config"], tab_id
        )
        if not page_config_queryset:
            queryset = []
        else:
            queryset = page_config_queryset.filter(filters).values(*keys)
        # 上传类控件数值变换
        queryset = value_to_list(upload_fields, queryset)
        logger.info("[get_list_components_data] get_querySet")
        if not need_page:
            return Response(self.get_response_data(queryset))
        pagination = CustomPageNumberPagination()
        page_queryset = pagination.paginate_queryset(queryset, self.request)
        return pagination.get_paginated_response(self.get_response_data(page_queryset))

    def get_page_config_queryset(self, manager, config, tab_id):
        """
        "show_mode": {
            "mode": 0,1,2
            "show_condition":[
                {
                    # 设置指定用户
                    "set_user_range": {
                        "members_group":[],
                        "members":[],
                        "condition": 0,1,2
                    },
                    # 展示相应用用户的数据
                    "display_user_range": {
                        "members_group":[],
                        "members":[]
                    },
                },
                ...
            ]
        }
        """
        config = self.get_config(config)
        queryset = manager.get_queryset()

        # 排序字段支持
        ordering = config.get("ordering", "id")
        # 排序字段支持
        if not isinstance(ordering, list):
            ordering = [ordering]
        queryset = queryset.order_by(*ordering)

        # 筛选条件支持
        if tab_id:
            # 子选项卡
            tab_config = config["tab_config"]
            conditions = {}
            for item in tab_config:
                if item["tab_id"] == tab_id:
                    conditions = config.get("conditions", {})
                    break
        else:
            # 默认选项卡
            conditions = config.get("conditions", {})
        filters = self.convert_conditions(conditions)

        if conditions:
            queryset = queryset.filter(filters)

        # all表示所有的数据
        time_range = config.get("time_range", "all")
        if time_range != "all":
            time_range_conditions = compute_time_range(time_range)
            queryset = queryset.filter(time_range_conditions)

        # 0:展示所有数据，1:展示用户自己的数据, 2:展示范围选择
        show_mode = config.get("show_mode", {})
        if not show_mode or show_mode["mode"] == 0:
            return queryset

        if show_mode["mode"] == 1:
            queryset = queryset.filter(creator=self.request.user.username)

        # 选择范围显示
        if show_mode["mode"] == 2:
            # username = self.request.user.username
            display_role = show_mode["display_role"].split(",")
            queryset = queryset.filter(creator__in=display_role)

        return queryset

    def export_list_component_data(self, conditions=None, ids=None):

        if not conditions:
            conditions = {}

        filters = self.convert_conditions(conditions)

        page_config = PageComponentHandler.get_list_component_config(
            self.page_id, self.version_number
        )
        worksheet_id = page_config["value"]
        manager = DataManager(worksheet_id)
        fields = manager.fields
        # 获取所要展示自定义字段
        keys, _ = self.get_keys(page_config["config"], fields)

        page_config_queryset = self.get_page_config_queryset(
            manager, page_config["config"], tab_id=None
        )
        # 导出批量记录
        if ids:
            base_queryset = page_config_queryset.filter(id__in=ids)
        else:
            base_queryset = page_config_queryset
        queryset = base_queryset.filter(filters).values(*keys)

        key_map = self.get_key_map(fields)
        # 初始化表格
        work_book = xlwt.Workbook(encoding="utf-8")
        work_sheet = work_book.add_sheet(manager.worksheet.name)
        # 表格样式初始化
        alignment = xlwt.Alignment()
        # 水平位置
        alignment.horz = xlwt.Alignment.HORZ_CENTER
        # 垂直方向
        alignment.vert = xlwt.Alignment.VERT_CENTER

        style = xlwt.XFStyle()
        # 样式加载
        style.alignment = alignment

        # 页面展示表格的字段作为首行内容
        head_fields = self.get_filter_fields(page_config["config"], fields)
        text_keys = [
            f"contents__{item['key']}" for item in head_fields if item["type"] == "TEXT"
        ]

        # 首行写入
        for index, value in enumerate(head_fields):
            work_sheet.col(index).width = 256 * 20
            work_sheet.write(0, index, value["name"])

        # 对应内容写入
        for row, values in enumerate(queryset):
            for index, key in enumerate(keys):
                if key in key_map:
                    field = key_map.get(key)

                    if field["type"] in ["SELECT", "MULTISELECT", "CHECKBOX", "RADIO"]:
                        choices = field.get("choice", [])

                        # 如果配置了数据源，优先使用数据源
                        if "data_config" in field.get("meta"):
                            value = values.get(key, "--")
                        elif choices:
                            try:
                                choice_keys = values.get(key, "--").split(",")
                            except Exception:
                                fields.append("")
                                continue

                            choice_map = {item["key"]: item["name"] for item in choices}
                            value = ""
                            for choice_key in choice_keys:
                                value += choice_map.get(choice_key, "") + ","
                            value = value.strip(",")
                        else:
                            value = values.get(key, "--")

                        work_sheet.write(row + 1, index, value, style)
                else:
                    if key in text_keys:
                        # 多行文本，自动换行
                        style.alignment.wrap = AUTO_LINE_FEED
                    work_sheet.write(row + 1, index, values.get(key, "--"), style)

        output = io.BytesIO()
        work_book.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(), content_type="application/vnd.ms-excel"
        )
        response["Content-Disposition"] = 'attachment;filename="{}.xls"'.format(
            manager.worksheet.name
        )

        return response

    def get_filter_fields(self, config, fields):
        filter_fields = [{"name": "id", "type": "INT"}]
        config = self.get_config(config)
        field_ids = config.get("fields", [])
        sys_fields = config.get("sys_fields", [])
        for field in fields:
            if field["id"] in field_ids:
                filter_fields.append(field)

        sys_enum = {
            "creator": {"name": "提交人", "type": "STRING"},
            "update_at": {"name": "更新时间", "type": "DATETIME"},
            "create_at": {"name": "创建时间", "type": "DATETIME"},
            "updated_by": {"name": "更新人", "type": "STRING"},
        }
        for sys_field in sys_fields:
            filter_fields.append(sys_enum.get(sys_field))
        return filter_fields

    def get_key_map(self, fields):
        key_map = {}  # {contents__id: choices}
        for field in fields:
            key = "contents__{}".format(field["key"])
            if field["type"] in [
                "SELECT",
                "INPUTSELECT",
                "MULTISELECT",
                "CHECKBOX",
                "RADIO",
            ]:
                key_map[key] = field

        return key_map

    def get_keys(self, config, fields):
        """
        计算出列表所需要展示的字段的key列表
        """
        keys = ["id"]

        config = self.get_config(config)
        field_ids = config.get("fields", [])
        upload_fields = []
        for field in fields:
            if field["id"] in field_ids:
                keys.append("contents__{}".format(field["key"]))
                # 图片控件类型
                #
                if field["type"] in ["IMAGE", "FILE"]:
                    upload_fields.append("contents__{}".format(field["key"]))

        # 增加返回系统字段
        sys_fields = config.get("sys_fields", [])
        if sys_fields:
            keys += sys_fields
        return keys, upload_fields

    def generate_export_template(self):

        page_config = PageComponentHandler.get_list_component_config(
            self.page_id, self.version_number
        )
        worksheet_id = page_config["value"]
        manager = DataManager(worksheet_id)
        fields = manager.fields

        sheet_name = (
            manager.worksheet.name
            + "_"
            + datetime.datetime.now().strftime("%Y%m%d%H%M")
        )
        work_book = openpyxl.Workbook()
        ws = work_book.worksheets[0]
        print(["{}({})".format(value["name"], value["key"]) for value in fields])
        ws.append(["{}({})".format(value["name"], value["key"]) for value in fields])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # response = HttpResponse(content_type='application/vnd.ms-excel')
        response["Content-Disposition"] = "attachment; filename={}.xlsx".format(
            sheet_name
        )
        work_book.save(response)
        return response


class ListComponentDetailHandler(BaseDataHandler):
    def __init__(self, service_id, pk, worksheet_id):
        self.service_id = service_id
        self.pk = pk
        self.worksheet_id = worksheet_id

    def data(self):
        try:
            keys, upload_keys = ServiceHandler(
                service_id=self.service_id, worksheet_id=self.worksheet_id
            ).get_keys()
            manager = DataManager(self.worksheet_id)

            queryset = manager.get_queryset()

            queryset = queryset.filter(id=self.pk).values(*keys)

            # 上传类控件数值变换
            queryset = value_to_list(upload_keys, queryset)

        except Exception as e:
            logger.info(
                "数据详情获取失败，service_id={}, worksheet_id={}, pk={},"
                "error={}".format(self.service_id, self.worksheet_id, self.pk, e)
            )
            raise GetDetailDataError()
        return self.get_response_data(queryset)


class WorkSheetDataHandler(BaseDataHandler):
    def __init__(self, worksheet_id, request):
        self.worksheet_id = worksheet_id
        self.request = request

    def queryset_unrepeated_not_null(self, fields, queryset):
        if len(fields) == 1 and queryset:
            key = "contents__{}".format(fields[0])
            exist_value = []
            new_queryset = []
            for item in queryset:
                if item[key] in exist_value or not item[key]:
                    continue
                else:
                    exist_value.append(item[key])
                    new_queryset.append(item)
            queryset = new_queryset

        return queryset

    def selected_type_field_escape(self, manager, fields, queryset):
        select_fields = {}
        for item in manager.fields:
            if item["key"] in fields and item["type"] in SELECT_FIELDS_TYPE:
                select_value = {}
                for value in item["choice"]:
                    select_value.setdefault(value["key"], value["name"])
                select_fields.setdefault(f"contents__{item['key']}", select_value)

        if select_fields:
            for item in queryset:
                for key, value in item.items():
                    if not select_fields.get(key):
                        continue
                    item[key] = select_fields[key][value]
        return queryset

    def data(self, conditions, fields, need_page):
        manager = DataManager(self.worksheet_id)
        keys = self.get_keys(fields)
        if not conditions["connector"] and not conditions["expressions"]:
            conditions = {}
        filters = self.convert_conditions(conditions)
        queryset = manager.get_queryset().filter(filters).values(*keys)

        # 只针对单一字段数据查询的时候后做过滤
        queryset = self.queryset_unrepeated_not_null(fields, queryset)
        # 选择类型字段转义
        queryset = self.selected_type_field_escape(manager, fields, queryset)

        if need_page:
            pagination = CustomPageNumberPagination()
            page_queryset = pagination.paginate_queryset(queryset, self.request)
            return pagination.get_paginated_response(
                self.get_response_data(page_queryset)
            )
        return self.get_response_data(queryset)

    def get_keys(self, fields):

        keys = ["id"]

        for field in fields:
            keys.append("contents__{}".format(field))
        return keys

    def get_fields_by_version(self, version_number):
        project_version = ProjectVersion.objects.get(version_number=version_number)
        worksheet_fields_version = project_version.worksheet_field[
            str(self.worksheet_id)
        ]

        data_struct = {}
        data_struct.setdefault(
            "fields", ["id", "create_at", "update_at", "creator", "updated_by"]
        )
        data_struct.setdefault("upload_fields", [])
        for field in worksheet_fields_version:
            data_struct["fields"].append("contents__{}".format(field["key"]))
            if field["type"] in ["IMAGE", "FILE"]:
                data_struct["upload_fields"].append("contents__{}".format(field["key"]))
        return data_struct

    def get_worksheet_data_by_version(self, version_number):
        keys = self.get_fields_by_version(version_number)
        manager = DataManager(self.worksheet_id)
        queryset = manager.get_queryset().values(*keys["fields"])

        if keys["upload_fields"]:
            for item in queryset:
                for field in keys["upload_fields"]:
                    if item[field]:
                        # 列表字符串转换成列表
                        try:
                            item[field] = ast.literal_eval(item[field])
                        except SyntaxError:
                            continue
                    else:
                        item[field] = []

        return self.get_response_data(queryset)


class ExportDataHandler:
    def __init__(self, worksheet_id, request):
        self.worksheet_id = worksheet_id
        self.pattern = re.compile(r"[(](.*?)[)]", re.S)
        self.request = request
        self.manager = DataManager(worksheet_id=self.worksheet_id)

    def get_keys(self, headers):
        keys = []
        version_fields = self.manager.fields
        version_field_key = [item["key"] for item in version_fields]
        for item in headers:
            item = str(item)
            key = self.pattern.findall(item)
            if len(key) != 1:
                raise ImportDataError(
                    "[ExportDataHandler][get_keys]数据导入失败，表头{}不符合规范".format(item)
                )
            if key[0] not in version_field_key:
                raise ImportDataError(
                    "[ExportDataHandler][get_keys]数据导入失败，表头{}不符合规范，表单已变更，请重新生成模板文件".format(
                        item
                    )
                )
            keys.append(key[0])
        return keys

    def validate(self, file):
        try:
            results = {}
            wb = openpyxl.load_workbook(file)
            sheet = wb.worksheets[0]
            rows_count = sheet.max_row  # 行数
            logger.info(
                "[ExportDataHandler][import_by_excels] 正在导入数据，{}条".format(rows_count)
            )
            if rows_count > 1000:
                raise ImportDataError()
            headers = next(sheet.values)
            keys = self.get_keys(headers=headers)

            for index, rowValues in enumerate(sheet.values):
                if index == 0:
                    continue
                data = {}
                for key, value in zip(keys, rowValues):
                    if isinstance(value, datetime.datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    data[key] = value
                try:
                    self.manager.serializer_validate(data)
                except Exception as e:
                    results[index] = str(e)
                    logger.info("数据校验失败，第{}行,错误原因 error={}".format(index, e))
            if results:
                raise DataValidateError("数据校验失败, 错误原因 error={}".format(results))
            return results
        except Exception as e:
            raise DataValidateError("数据校验失败, 错误原因 error={}".format(e))

    def import_by_excels(self, file):
        wb = openpyxl.load_workbook(file)
        sheet = wb.worksheets[0]
        logger.info("[ExportDataHandler][import_by_excels] 正在校验数据的合法性")
        # 提前将所有数据的合法性校验一编
        rows_count = sheet.max_row  # 行数
        logger.info(
            "[ExportDataHandler][import_by_excels] 正在导入数据，{}条".format(rows_count)
        )
        headers = next(sheet.values)
        keys = self.get_keys(headers=headers)
        for index, rowValues in enumerate(sheet.values):
            if index == 0:
                continue
            data = {}
            for key, value in zip(keys, rowValues):
                if isinstance(value, datetime.datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                data[key] = value
            try:
                self.manager.add(data, self.request.user.username)
            except Exception as e:
                logger.info(
                    "[ExportDataHandler][import_by_excels] 数据导入失败，{}条, error={}".format(
                        rows_count, e
                    )
                )
                raise ImportDataError(
                    "[ExportDataHandler][import_by_excels] 数据导入失败，成功导入{}条".format(index)
                )
        return rows_count


class ChartDataHandler(ListComponentDataHandler):
    """
    {
        ...
        config：{
            name：图标名称,
            value：图表的类型,
        xaxes：{
            "key": field_key/create_at/update_at
            "type": const/time
            "value": (day/month/year)
            [value只针对create_at/update_at]
        },
        y_fields_list:  [
            {"type":}
        ],

        conditions:{
             expressions:[
                    {
                      "key": "create_at",
                      "type": "const",
                      "condition": "in",
                      "value": "1",
                    },
                    ...
                ]
        },

        }
    }

    """

    def __init__(
        self, request, page_component_id=None, page_id=None, version_number=None
    ):
        self.page_component_id = page_component_id
        self.condition = CONDITIONS_PROTOCOL
        super(ChartDataHandler, self).__init__(
            page_id=page_id, request=request, version_number=version_number
        )

    def analysis(self, chart_configs):
        result = self.analysis_data(chart_configs)
        return result

    def get_chart_config(self):
        # 获取该页面下的所有图表组件的设置
        chart_configs = PageComponentHandler.get_chart_component_config(
            page_id=self.page_id,
            version_number=self.version_number,
            page_component_id=self.page_component_id,
        )
        if isinstance(chart_configs, str):
            chart_configs = json.loads(chart_configs)
        return chart_configs

    @property
    def current_week(self):
        now = datetime.datetime.now()
        this_week_start = now - datetime.timedelta(days=now.weekday())
        this_week_end = now + datetime.timedelta(days=6 - now.weekday())
        time_range = (this_week_start, this_week_end)
        return time_range

    @property
    def current_month(self):
        now = datetime.datetime.now()
        this_month_start = datetime.datetime(now.year, now.month, 1)
        this_month_end = (
            datetime.datetime(now.year, now.month + 1, 1)
            - datetime.timedelta(days=1)
            + datetime.timedelta(hours=23, minutes=59, seconds=59)
        )
        time_range = (this_month_start, this_month_end)
        return time_range

    @property
    def current_year(self):
        now = datetime.datetime.now()
        this_year_start = datetime.datetime(now.year, 1, 1).strftime(
            "%Y-%m-%d %H:%M:%S.%f"
        )
        this_year_end = (
            datetime.datetime(now.year + 1, 1, 1)
            - datetime.timedelta(days=1)
            + datetime.timedelta(hours=23, minutes=59, seconds=59)
        )
        time_range = (this_year_start, this_year_end.strftime("%Y-%m-%d %H:%M:%S.%f"))
        return time_range

    @property
    def time_selector(self):
        return {
            "current_week": self.current_week,
            "current_month": self.current_month,
            "current_year": self.current_year,
        }

    def time_filter(self, time_range):
        if time_range["type"].upper() == "DEFINE":
            return time_range["conditions"]
        time_range = self.time_selector[time_range["type"].lower()]
        (
            self.condition["expressions"][0]["value"],
            self.condition["expressions"][1]["value"],
        ) = (time_range[0], time_range[1])
        return self.condition

    def analysis_data(self, chart_configs):
        all_chart = []
        for chart_config in chart_configs:
            # 获取chart组件设置
            chart_setting = chart_config
            worksheet_id = chart_config["worksheet_id"]
            time_range = chart_setting["time_range"]

            # 数据调取
            manager = DataManager(worksheet_id)
            # 时间范围过滤条件
            time_conditions = self.time_filter(time_range)
            chart_setting["time_range"]["conditions"] = time_conditions
            time_filters = self.convert_conditions(time_conditions)
            # 时间范围过滤后所有数据
            queryset = manager.get_queryset().filter(time_filters)

            # 获取X轴过滤依据
            """
            {
                "key":field_key,
                "type": const/time,
                "value": (只针对type=time)
            ]
            """
            x_label_key = chart_setting["xaxes"]

            # 获取Y轴依据
            y_label_list = chart_setting["yaxis_list"]

            # all contents
            if x_label_key["type"] == "const":
                result = self.analysis_by_const(
                    queryset=queryset,
                    x_label_key=x_label_key,
                    y_label_list=y_label_list,
                    chart_setting=chart_setting,
                )
                all_chart.append(result)

            if x_label_key["type"] == "time":
                result = self.analysis_by_time(
                    queryset=queryset,
                    x_label_key=x_label_key,
                    y_label_list=y_label_list,
                    chart_setting=chart_setting,
                )
                all_chart.append(result)

        return all_chart

    def count(self, content_list, depend_field, depend_value):
        if depend_field == "total":
            return len(content_list)
        count = 0
        for item in content_list:
            if item.get(depend_field) == depend_value:
                count += 1
        return count

    def sum(self, content_list, depend_field):
        value = 0
        for content in content_list:
            try:
                value += int(content.get(depend_field))
            except TypeError:
                continue
        return value

    def analysis_by_time(self, queryset, x_label_key, y_label_list, chart_setting):
        """
        {
            "key":field_key,
            "type": const/time,
            "value": (只针对type=time)
        }
        """
        if x_label_key["key"] == "create_at":
            queryset.order_by("-create_at")
        if x_label_key["key"] == "update_at":
            queryset.order_by("-update_at")

        # 获取时间范围
        year_filter_1 = int(
            chart_setting["time_range"]["conditions"]["expressions"][0]["value"].split(
                "-"
            )[0]
        )
        year_filter_2 = int(
            chart_setting["time_range"]["conditions"]["expressions"][1]["value"].split(
                "-"
            )[0]
        )
        # 过滤年生成器
        first_params = min(year_filter_2, year_filter_1)
        second_prams = max(year_filter_2, year_filter_1)
        year_range = (
            range(first_params, second_prams + 1)
            if first_params != second_prams
            else [first_params]
        )

        sorted_content = self.sort_queryset_by_time(
            queryset=queryset,
            x_label_key=x_label_key,
            year_range=year_range,
            chart_setting=chart_setting,
        )

        y_result = self.analysis_after_sort_content(
            sorted_content=sorted_content, y_label_list=y_label_list
        )

        x_name = x_label_key["name"]
        y_name = y_label_list[0]["name"]
        result = []
        for key, value in y_result.items():
            result.append({x_name: key, y_name: value[0]["result"]})
        return result

    def analysis_by_const(self, queryset, x_label_key, y_label_list, chart_setting):
        # x轴分类的数据
        sorted_content = self.sort_queryset_by_key(queryset, x_label_key)
        # 根据y轴依据进行数据的统计
        y_result = self.analysis_after_sort_content(sorted_content, y_label_list)

        x_name = x_label_key["name"]
        y_name = y_label_list[0]["name"]
        result = []
        for key, value in y_result.items():
            result.append({x_name: key, y_name: value[0]["result"]})
        return result

    def analysis_after_sort_content(self, sorted_content, y_label_list):
        y_result = {}
        for field, content_list in sorted_content.items():
            y_result[field] = []
            for y_label in y_label_list:
                # 创建分类的总数统计
                if y_label["type"] == "count":
                    result_data = self.count(
                        content_list, y_label["field"], y_label["value"]
                    )
                # 依据字段总计统计
                # if y_label["type"] == "sum":
                else:
                    result_data = self.sum(content_list, y_label["field"])
                result_data = {
                    "field_key": y_label["field"],
                    "field_value": y_label["value"],
                    "analysis_type": y_label["type"],
                    "result": result_data,
                }

                y_result[field].append(result_data)

        return y_result

    def sort_queryset_by_key(self, queryset, x_label_key):
        x_sort = []
        sorted_content = {}

        for item in queryset:
            content = (
                json.loads(item.contents)
                if isinstance(item.contents, str)
                else item.contents
            )
            # x轴 分类 同时根据x轴分类数据
            # 根据key，获取分类
            sort_value = content.get(x_label_key["key"])
            # 数据分类
            if sort_value in x_sort:
                sorted_content[sort_value].append(content)
            else:
                x_sort.append(sort_value)
                sorted_content.setdefault(sort_value, [content])
        return sorted_content

    def sort_queryset_by_time(self, queryset, x_label_key, year_range, chart_setting):
        """
        {
            "key":field_key,
            "type": const/time,
            "value": (只针对type=time)
        }
        """
        filter_condition = x_label_key["value"]
        # 按照天的分类
        if filter_condition == DAY:
            return self.sort_queryset_by_day(
                queryset=queryset,
                x_label_key=x_label_key,
                year_range=year_range,
                chart_setting=chart_setting,
            )

        # 按照每年月份的分类
        if filter_condition == MONTH:
            # 根据年月对content 进行分类
            return self.sort_queryset_by_month(
                queryset=queryset,
                x_label_key=x_label_key,
                year_range=year_range,
            )

        # 按照年分类
        if filter_condition == YEAR:
            return self.sort_queryset_by_year(
                queryset=queryset, x_label_key=x_label_key, year_range=year_range
            )

    def sort_queryset_by_day(self, queryset, x_label_key, year_range, chart_setting):
        x_sort = []
        sorted_content = {}
        for year in year_range:
            # 月
            month = chart_setting["time_range"]["conditions"]["expressions"][0][
                "value"
            ].split("-")[1]
            each_day_queryset = None
            if x_label_key["key"] == "create_at":
                filter_queryset = queryset.filter(
                    create_at__year=year, create_at__month=month
                )
                each_day_queryset = (
                    filter_queryset.extra(
                        select={"create_at": "DATE_FORMAT(create_at, '%%e')"}
                    )
                    .values("create_at")
                    .annotate(total=Count("create_at"))
                    .values("create_at", "total", "contents")
                )

            if x_label_key["key"] == "update_at":
                filter_queryset = queryset.filter(
                    update_at__year=year, update_at__month=month
                )
                each_day_queryset = (
                    filter_queryset.extra(
                        select={"update_at": "DATE_FORMAT(create_at, '%%e')"}
                    )
                    .values("update_at")
                    .annotate(total=Count("update_at"))
                    .values("update_at", "total", "contents")
                    .order_by("-create_at")
                )
            for item in each_day_queryset:
                create_at_day = item.get("create_at")
                sort_value = f"{year}-{month}-{create_at_day}"

                if sort_value in x_sort:
                    sorted_content[sort_value].append(item.get("contents"))
                else:
                    x_sort.append(sort_value)
                    sorted_content.setdefault(sort_value, [item.get("contents")])
        return sorted_content

    def sort_queryset_by_month(self, queryset, x_label_key, year_range):
        x_sort = []
        sorted_content = {}
        for year in year_range:
            for month in range(1, 13):
                sort_value = f"{year}-{month}"
                if x_label_key["key"] == "create_at":
                    each_month_queryset = queryset.filter(
                        create_at__year=year, create_at__month=month
                    )
                if x_label_key["key"] == "update_at":
                    each_month_queryset = queryset.filter(
                        update_at__year=sort_value, update_at__month=month
                    )

                for item in each_month_queryset:
                    content = (
                        json.loads(item.contents)
                        if isinstance(item.contents, str)
                        else item.contents
                    )

                    # 2021-10-1
                    if sort_value in x_sort:
                        sorted_content[sort_value].append(content)
                    else:
                        x_sort.append(sort_value)
                        sorted_content.setdefault(sort_value, [content])
        return sorted_content

    def sort_queryset_by_year(self, queryset, x_label_key, year_range):
        x_sort = []
        sorted_content = {}

        for year in year_range:
            sort_value = f"{year}"
            if x_label_key["key"] == "create_at":
                # 时间范围过滤后所有数据
                each_year_queryset = queryset.filter(create_at__year=sort_value)
            if x_label_key["key"] == "update_at":
                # 时间范围过滤后所有数据
                each_year_queryset = queryset.filter(update_at__year=sort_value)

            for item in each_year_queryset:
                content = (
                    json.loads(item.contents)
                    if isinstance(item.contents, str)
                    else item.contents
                )

                # 2021-10-1
                if sort_value in x_sort:
                    sorted_content[sort_value].append(content)
                else:
                    x_sort.append(sort_value)
                    sorted_content.setdefault(sort_value, [content])
        return sorted_content
