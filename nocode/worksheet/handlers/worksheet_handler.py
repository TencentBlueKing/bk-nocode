# -*- coding: utf-8 -*-
"""
Tencent is pleased to support the open source community by making BK-NOCODE SMAKER蓝鲸无代码平台  available.

Copyright (C) 2021 THL A29 Limited, a Tencent company.  All rights reserved.

BK-NOCODE 蓝鲸无代码平台(S-maker) is licensed under the MIT License.

License for BK-NOCODE 蓝鲸无代码平台(S-maker) :
--------------------------------------------------------------------
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation
the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software,
and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial
portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT
LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN
NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY,
WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""
import datetime
import uuid

import openpyxl
from common.log import logger

from itsm.service.handler.service_handler import ServiceHandler as ServiceModelHandler
from nocode.base.base_handler import APIModel
from nocode.base.constants import WORKSHEET
from nocode.page.handlers.page_handler import PageComponentHandler, PageModelHandler
from nocode.project_manager.handlers.project_white_handler import (
    ProjectWhiteHandler,
)
from nocode.project_manager.models import ProjectWhite
from nocode.worksheet.exceptions import (
    WorkSheetDoesNotExist,
    NotSupportExcelFileType,
    InitWorkSheetError,
)
from nocode.worksheet.handlers.project_version_handler import ProjectVersionHandler
from nocode.worksheet.models import WorkSheet


class WorkSheetModelHandler(APIModel):
    def __init__(self, worksheet_id=None):
        self.worksheet_id = worksheet_id
        self.obj = None
        super(WorkSheetModelHandler, self).__init__()

    def _get_instance(self):
        try:
            obj = WorkSheet._objects.get(id=self.worksheet_id)
        except WorkSheet.DoesNotExist:
            raise WorkSheetDoesNotExist("没有工作表")
        return obj

    @property
    def all_worksheet(self):
        return WorkSheet.objects.all()

    def filter(self, *args, **kwargs):
        return self.all_worksheet.filter(*args, **kwargs)

    @property
    def instance(self):
        return self._get_instance()

    def update_fields(self, fields):
        if self.obj is None:
            self.obj = self._get_instance()
        self.obj.fields = fields
        self.obj.save()

    @property
    def fields(self):
        if self.worksheet_id is not None:
            return ProjectVersionHandler(self.instance.project_key).get_fields(
                str(self.worksheet_id)
            )
        raise WorkSheetDoesNotExist("worksheet_id is None")

    def delete_service(self):
        from nocode.worksheet.handlers.moudule_handler import ServiceHandler

        ServiceHandler(self.instance).delete_service()

    def get_worksheet_relate_service_page(self, worksheet):
        worksheet_struct_data = dict()
        # 数据初始化
        worksheet_struct_data.update(worksheet.tag_data())
        worksheet_struct_data.setdefault("relate_service", [])
        worksheet_struct_data.setdefault("relate_list_page", [])

        # 绑定了表单的列表组件
        list_page_ids = (
            PageComponentHandler()
            .filter(value=worksheet.id, type="LIST")
            .values_list("page_id", flat=True)
        )
        # 应用下的功能
        project_of_services = (
            ServiceModelHandler()
            .filter(
                project_key=worksheet.project_key,
                is_deleted=False,
            )
            .values("id", "name", "is_builtin", "worksheet_ids")
        )
        # 表单绑定的功能
        for item in project_of_services:
            service_worksheet_ids = item.pop("worksheet_ids")
            if worksheet.id in service_worksheet_ids:
                worksheet_struct_data["relate_service"].append(item)
        # 表单绑定的页面
        for list_page in (
            PageModelHandler().filter(id__in=list_page_ids).values("id", "name")
        ):
            item = {"id": list_page["id"], "name": list_page["name"]}
            worksheet_struct_data["relate_list_page"].append(item)
        return worksheet_struct_data

    def get_all_worksheet_relate_service_page(self, project_key):
        worksheet_struct_data = dict()
        for instance in self.filter(project_key=project_key):
            worksheet_data = instance.tag_data()
            worksheet_data.setdefault("relate_service", [])
            worksheet_data.setdefault("relate_list_page", [])
            worksheet_struct_data.setdefault(worksheet_data.get("id"), worksheet_data)

        project_of_services = (
            ServiceModelHandler()
            .filter(
                project_key=project_key,
                is_deleted=False,
            )
            .values("id", "name", "is_builtin", "worksheet_ids")
        )
        list_pages = (
            PageModelHandler()
            .filter(type="LIST", project_key=project_key, is_deleted=False)
            .values("id", "name")
        )

        all_worksheet_ids = worksheet_struct_data.keys()

        for item in project_of_services:
            service_worksheet_ids = item.pop("worksheet_ids")
            for worksheet_id in all_worksheet_ids:
                if worksheet_id in service_worksheet_ids:
                    worksheet_struct_data[worksheet_id]["relate_service"].append(item)

        for list_page in list_pages:
            component = (
                PageComponentHandler()
                .filter(page_id=list_page["id"], is_deleted=False)
                .first()
            )

            item = {"id": list_page["id"], "name": list_page["name"]}

            worksheet_struct_data[int(component.value)]["relate_list_page"].append(item)

        return worksheet_struct_data.values()

    def delete_worksheets(self, project_key):
        return WorkSheet._objects.filter(is_deleted=True, project_key=project_key)

    def init_white_project(self):
        data = {
            "value": self.instance.id,
            "project_key": self.instance.project_key,
            "type": "WORKSHEET",
            "granted_project": {"projects": []},
        }
        ProjectWhiteHandler.init_white_project(data)

    def get_white_project(self):
        try:
            white_project = ProjectWhite.objects.get(
                value=self.instance.id, type=WORKSHEET
            ).granted_project
            return white_project["projects"]
        except ProjectWhite.DoesNotExist:
            logger.info("worksheet_id {}, 没有设置白名单".format(self.instance.id))
            return []


class InitWorksheetHandler(object):
    def __init__(self, file, sheet_name):
        self.file = file
        self.sheet_name = sheet_name
        excel_type = file.name.split(".")[1]
        if excel_type not in ["xls", "xlsx"]:
            raise NotSupportExcelFileType("导入文件不符合规范")

    def get_choice_field(self, ws, data):
        type_map = {}
        validations = ws.data_validations.dataValidation
        for validation in validations:
            ranges = validation.sqref.ranges
            if len(ranges) != 1:
                continue
            if validation.type == "list":
                list_cells = ws[validation.formula1]
                values = [cell.value for cell_row in list_cells for cell in cell_row]
                bounds = ranges[0].bounds
                index = data[bounds[0] - 1]
                type_map[index] = values
        return type_map

    def get_type(self, value):
        if isinstance(value, int):
            return "INT"
        if isinstance(value, str):
            return "STRING"
        if isinstance(value, float):
            return "STRING"
        if isinstance(value, datetime.datetime):
            return "DATE"

        return "STRING"

    def get_filed_type(self, data, choice_map):
        type_map = []
        for index, value in zip(data[0], data[1]):
            if index in choice_map:
                continue
            type_map.append({"name": index, "type": self.get_type(value)})
        for key, value in choice_map.items():
            type_map.append(
                {
                    "name": key,
                    "type": "SELECT",
                    "choice": [
                        {"key": uuid.uuid4().hex, "name": item} for item in value
                    ],
                }
            )
        return type_map

    def build_base_field_type(self, data, choice_map):
        type_map = []
        for index in data[0]:
            if index in choice_map:
                continue
            type_map.append({"name": index, "type": "STRING"})
        for key, value in choice_map.items():
            type_map.append(
                {
                    "name": key,
                    "type": "SELECT",
                    "choice": [
                        {"key": uuid.uuid4().hex, "name": item} for item in value
                    ],
                }
            )
        return type_map

    def get_init_filed_details(self):
        """
        会选择前两行数据，综合去判断字段类型是什么
        """
        wb = openpyxl.load_workbook(self.file)
        try:
            ws = wb[self.sheet_name]
        except Exception:
            ws = wb.worksheets[0]

        if ws.max_row < 1:
            raise InitWorkSheetError("初始化工作表失败，请确保拥有1行数据")

        data = []
        for item in ws.values:
            data.append(item)
            if len(data) == 2:
                break
        choice_map = self.get_choice_field(ws, data[0])

        if ws.max_row > 2:
            filed_type_map = self.get_filed_type(data, choice_map)

        if ws.max_row == 1:
            filed_type_map = self.build_base_field_type(data, choice_map)

        return filed_type_map
